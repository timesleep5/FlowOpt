from typing import List

from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter, get_column_interval, column_index_from_string


class Workbookutils:
    # Filling
    @staticmethod
    def fill_row(ws, startcol: str, row: int, content: List[str]):
        col_interval = get_column_interval(startcol,
                                              get_column_letter(column_index_from_string(startcol) + len(content) - 1))
        for col, cell_content in zip(col_interval, content):
            ws[f'{col}{row}'] = cell_content  # to be at Index 0 in the first iteration

    @staticmethod
    def fill_col(ws, startcell: str, content: List[str]):
        col, startrow = coordinate_from_string(startcell)
        rows = [row for row in range(int(startrow), int(startrow) + len(content))]
        for row, cell_content in zip(rows, content):
            ws[f'{col}{row}'] = cell_content

    # Removing
    @staticmethod
    def remove_items_from_col(ws, col: str, endrow: int, items: List, way: str):
        if way == 'name':
            for item in items:
                for i in range(1, endrow):
                    if ws[f'{col}{i}'].value == item:
                        ws[f'{col}{i}'].value = None
        elif way == 'index':
            for item in items:
                ws[f'{col}{item}'].value = None

    # Styling (bold, italic, ...)
    @staticmethod
    def style_row_font(ws, startcol, endcol, row, style):
        for i in range(startcol, endcol + 1):
            char = get_column_letter(i)
            if style == 'bold':
                ws[f'{char}{row}'].font = Font(bold=True)
            elif style == 'italic':
                ws[f'{char}{row}'].font = Font(italic=True)

    @staticmethod
    def style_col_font(ws, startrow, endrow, col, style):
        char = get_column_letter(col)
        for i in range(startrow, endrow + 1):
            if style == 'bold':
                ws[f'{char}{i}'].font = Font(bold=True)
            elif style == 'italic':
                ws[f'{char}{i}'].font = Font(italic=True)

    # Border Methods
    @staticmethod
    def _create_border(shape, style):
        side = Side(style=style)
        if shape == 'full':
            return Border(top=side, left=side, bottom=side, right=side)
        elif shape == 'sides':
            return Border(left=side, right=side)
        elif shape == 'opentop':
            return Border(left=side, right=side, bottom=side)

    @staticmethod
    def _apply_border(ws, cell, shape, style):
        border = Workbookutils._create_border(shape, style)
        ws[cell].border = border

    @staticmethod
    def style_row_border(ws, startcol: int, endcol: int, row: int, shape: str, style: str):
        for i in range(startcol, endcol + 1):
            char = get_column_letter(i)
            Workbookutils._apply_border(ws=ws, cell=f'{char}{row}', shape=shape, style=style)

    @staticmethod
    def style_row_border_mult(ws, startcol: int, endcol: int, startrow: int, endrow: int, shape: str, style: str):
        for i in range(startcol, endcol + 1):
            for j in range(startrow, endrow + 1):
                char = get_column_letter(i)
                Workbookutils._apply_border(ws=ws, cell=f'{char}{j}', shape=shape, style=style)

    @staticmethod
    def frame_multiple_cells(ws, range: str):
        pass

    # Custom Methods
    @staticmethod
    def find_next_empty_col_cell(ws, col: str):
        for row in range(1, 30):
            if ws[f'{col}{row}'].value is None:
                return row

    @staticmethod
    def reorder_column(ws, col: str, endrow):
        startrow = Workbookutils.find_next_empty_col_cell(ws, col)
        items_to_reorder = Workbookutils.get_items_to_reorder(ws, col, startrow, endrow)
        startcell = col + str(startrow)
        Workbookutils.fill_col(ws, startcell, items_to_reorder)

    @staticmethod
    def get_items_to_reorder(ws, col: str, startrow: int, endrow: int) -> List[str]:
        items_to_reorder = []

        for row in range(startrow, endrow + 1):
            if ws[f'{col}{row}'].value is not None:
                items_to_reorder.append(ws[f'{col}{row}'].value)
                ws[f'{col}{row}'].value = None
        return items_to_reorder

    @staticmethod
    def get_names_by_index(ws, col: str, items: List[int]) -> List[str]:
        named_items = []

        for row in items:
            named_items.append(ws[f'{col}{row}'].value)

        return named_items

    @staticmethod
    def get_items_in_space(ws, space: str) -> List[List[str]]:
        startcol, startrow, _, endcol, endrow = space
        col_interval = get_column_interval(startcol, endcol)
        items = []
        for char in col_interval:
            col_content = []
            for row in range(int(startrow), int(endrow) + 1):
                if ws[f'{char}{row}'].value is None:
                    continue
                col_content.append(ws[f'{char}{row}'].value)
            items.append(col_content)

        return items

    @staticmethod
    def fill_items_in_space(ws, startpt: str, items: List[List[str]]):
        startcol, startrow = coordinate_from_string(startpt)
        col_interval = get_column_interval(startcol,
                                           get_column_letter(column_index_from_string(startcol) + len(items) - 1))
        for col, col_content in zip(col_interval, items):
            row = startrow
            for cell_content in col_content:
                ws[f'{col}{row}'] = cell_content
                row += 1

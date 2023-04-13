from datetime import date
from typing import List

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from Utils.simpledatautils import SimpleDataUtils as sd_utils

from Utils.workbookutils import Workbookutils as w_utils

from Utils.timeutils import Timeutils as t_utils


class ChangeWorkbook:
    def __init__(self, wbname):
        self.wb = self.get_workbook(wbname)
        self.wbname = wbname
        self.week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        self.categories = ['Day', 'Date', 'Planned', 'Done', 'Resumé']
        self._insert_history_table()

    @staticmethod
    def get_workbook(wbname):
        return load_workbook(wbname)

    def update_dates(self, wsname: str, startcell: str):
        ws = self.wb[wsname]
        dates = t_utils.get_week_dates(date.today(), 1, 7)
        startchar, row = startcell
        w_utils.fill_row(ws, startchar, row, dates)

    def add_to_table(self, ws: str, col: str, items: List[str]):
        ws = self.wb[ws]
        next_line = w_utils.find_next_empty_col_cell(ws, col)
        startcell = col + str(next_line)
        w_utils.fill_col(ws, startcell, items)

    def remove_from_table(self, wsname: str, col: str, items: List, way: str):
        ws = self.wb[wsname]
        endrow = w_utils.find_next_empty_col_cell(ws, col)

        w_utils.remove_items_from_col(ws, col, endrow, items, way)
        w_utils.reorder_column(ws, col, 30) # bis dahin soll die Funktion nach den zu löschenden Wörtern suchen

    def move_from_backlog_to_schedule(self, col: str, items: List):
        ws = self.wb['Backlog']
        if type(items[0]) == int:
            items = w_utils.get_names_by_index(ws, col, items)
        self.remove_from_table('Backlog', col, items, 'name')

        ws = self.wb['Schedule']
        startrow = w_utils.find_next_empty_col_cell(ws, col)
        startcell = col + str(startrow)
        w_utils.fill_col(ws, startcell, items)

    def move_from_schedule_to_backlog(self, col: str, items: List):
        ws = self.wb['Schedule']
        if type(items[0]) == int:
            items = w_utils.get_names_by_index(ws, col, items)
        self.remove_from_table('Schedule', col, items, 'name')

        ws = self.wb['Backlog']
        startrow = w_utils.find_next_empty_col_cell(ws, col)
        startcell = col + str(startrow)
        w_utils.fill_col(ws, startcell, items)

    def clear_schedule(self):
        ws = self.wb['Schedule']
        col_range = [x for x in range(4, 9)]
        for i in range(1, 8):
            col = get_column_letter(i)
            w_utils.remove_items_from_col(ws, col, 8, col_range, 'index')

    def _insert_history_table(self):
        ws = self.wb['History']
        w_utils.fill_row(ws, 'B', 1, self.week)
        w_utils.fill_col(ws, 'A1', self.categories)
        w_utils.style_col_font(ws, 1, len(self.categories), 1, 'bold')

    def add_new_history_element(self):
        ws = self.wb['History']
        for _ in range(7):
            ws.insert_rows(1)
        self._insert_history_table()

    def update_history(self):
        self.add_new_history_element()

        ws = self.wb['Schedule']
        items = w_utils.get_items_in_space(ws, 'A4:G8')
        longest_list = sd_utils.find_longest_list_length(items)

        ws = self.wb['History']
        self.update_dates('History', 'B2')
        for _ in range(longest_list-1):
            ws.insert_rows(4)
        w_utils.fill_items_in_space(ws, 'B3', items)

    def savewb(self, name):
        self.wb.save(filename=name)
